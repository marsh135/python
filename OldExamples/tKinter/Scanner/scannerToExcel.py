import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from tkinter import simpledialog
import openpyxl

class StudentManager:
    def __init__(self):
        self.students = {}

    def add_student(self, student_id, name):
        if student_id not in self.students:
            self.students[student_id] = {'name': name, 'checked_in': False}
            return f"Student {name} with ID {student_id} added successfully."
        else:
            return f"Student {name} with ID {student_id} already exists."

    def get_student_name(self, student_id):
        return self.students.get(student_id, {}).get('name', None)

    def get_student_checked_in_status(self, student_id):
        return self.students.get(student_id, {}).get('checked_in', False)

    def log_event(self, student_id, action):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        event = f"{timestamp} - Student {student_id} {action}"
        with open("student_log.txt", "a") as log_file:
            log_file.write(event + "\n")

        # Log to Excel spreadsheet
        workbook = openpyxl.load_workbook("student_log.xlsx")
        sheet = workbook.active
        sheet.append([timestamp, student_id, action])
        workbook.save("student_log.xlsx")

class StudentScannerGUI:
    def __init__(self, root, student_manager):
        self.root = root
        self.root.title("Student Scanner")
        self.student_manager = student_manager

        self.label_action = tk.Label(root, text="Scan 'in' or 'out':")
        self.label_action.pack()

        self.entry_action = tk.Entry(root)
        self.entry_action.pack()

        self.label_id = tk.Label(root, text="Scan student ID:")
        self.label_id.pack()

        self.entry_id = tk.Entry(root)
        self.entry_id.pack()

        self.btn_scan = tk.Button(root, text="Scan and Confirm", command=self.scan_and_confirm)
        self.btn_scan.pack()

        self.btn_add_student = tk.Button(root, text="Add Student", command=self.add_student)
        self.btn_add_student.pack()

        # Bind the entry widget to a callback function for barcode scanning
        self.entry_id.bind("<KeyPress>", self.on_entry_keypress)

    def scan_and_confirm(self):
        action = self.entry_action.get().lower()
        student_id = self.entry_id.get()

        if action == 'in' or action == 'out':
            if self.student_manager.get_student_name(student_id):
                if action == 'in':
                    self.check_in(student_id)
                elif action == 'out':
                    self.check_out(student_id)
            else:
                messagebox.showinfo("Error", f"Student with ID {student_id} not found.")
        else:
            messagebox.showinfo("Error", "Invalid action. Please enter 'in' or 'out.'")

    def check_in(self, student_id):
        if self.student_manager.get_student_checked_in_status(student_id):
            messagebox.showinfo("Error", f"Student {student_id} is already checked in.")
        else:
            self.student_manager.students[student_id]['checked_in'] = True
            self.student_manager.log_event(student_id, "checked in")
            messagebox.showinfo("Success", f"Student {student_id} checked in successfully.")

    def check_out(self, student_id):
        if self.student_manager.get_student_checked_in_status(student_id):
            self.student_manager.students[student_id]['checked_in'] = False
            self.student_manager.log_event(student_id, "checked out")
            messagebox.showinfo("Success", f"Student {student_id} checked out successfully.")
        else:
            messagebox.showinfo("Error", f"Student {student_id} is already checked out.")

    def add_student(self):
        student_id = self.entry_id.get()
        name = simpledialog.askstring("Input", "Enter student name:")
        if name:
            result = self.student_manager.add_student(student_id, name)
            messagebox.showinfo("Result", result)

    def on_entry_keypress(self, event):
        # Check if the Enter key is pressed (barcode scanner usually sends Enter after scanning)
        if event.keysym == "Return":
            self.scan_and_confirm()

if __name__ == "__main__":
    try:
        workbook = openpyxl.load_workbook("student_log.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Timestamp", "Student ID", "Action"])
        workbook.save("student_log.xlsx")

    student_manager = StudentManager()

    root = tk.Tk()
    app = StudentScannerGUI(root, student_manager)
    root.mainloop()
